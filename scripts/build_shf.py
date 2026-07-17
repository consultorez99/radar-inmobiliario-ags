#!/usr/bin/env python3
"""
Genera data/ags_shf_indice.json: serie trimestral del Índice SHF de Precios
de la Vivienda (base 2017=100) para Nacional, estado de Aguascalientes y los
municipios de Aguascalientes y Jesús María, desde 1T2005.

Dato abierto oficial (Libre Uso MX) de Sociedad Hipotecaria Federal — a
diferencia del estudio de mercado de terceros, esta serie sí se puede
versionar y citar libremente.

Insumo:  data/raw/shf/Indice_SHF_datos_abiertos_*_trim_*.xlsx
         Descarga trimestral desde https://www.gob.mx/shf/documentos/90727
         (artículo "Índice SHF de Precios de la Vivienda en México" del
         trimestre más reciente → liga "Datos Abiertos").

Uso:  .venv/bin/python scripts/build_shf.py   (requiere openpyxl)
"""

import glob
import json
import os

from openpyxl import load_workbook

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RAW_GLOB = os.path.join(BASE, "data/raw/shf/Indice_SHF_datos_abiertos_*.xlsx")
OUT = os.path.join(BASE, "data/ags_shf_indice.json")

# (columna del XLSX que identifica la serie, valor a buscar, clave y nombre de salida)
SERIES = [
    ("Global", "Nacional", "nacional", "Nacional"),
    ("Estado", "Aguascalientes", "estado", "Aguascalientes (estado)"),
    ("Municipio", "Aguascalientes", "municipio_ags", "Aguascalientes (municipio)"),
    ("Municipio", "Jesús María", "jesus_maria", "Jesús María (municipio)"),
]


def main():
    candidatos = sorted(glob.glob(RAW_GLOB))
    if not candidatos:
        raise SystemExit(f"No hay insumo. Descarga el XLSX de datos abiertos a {RAW_GLOB}")
    ruta = candidatos[-1]  # el más reciente por orden alfabético del nombre
    print(f"Insumo: {ruta}")

    ws = load_workbook(ruta, read_only=True, data_only=True).active
    filas = ws.iter_rows(values_only=True)
    encabezado = [str(c).strip() if c else "" for c in next(filas)]
    col = {nombre: encabezado.index(nombre) for nombre in
           ("Global", "Estado", "Municipio", "Trimestre", "Año", "Indice")}

    datos = {clave: [] for _, _, clave, _ in SERIES}
    for fila in filas:
        if fila[col["Indice"]] is None:
            continue
        celda = {c: (str(fila[i]).strip() if fila[i] is not None else "") for c, i in col.items()}
        for columna, valor, clave, _ in SERIES:
            # Estado sin municipio = serie estatal; municipio pide además el estado correcto
            if celda[columna] != valor:
                continue
            if columna == "Estado" and celda["Municipio"]:
                continue
            if columna == "Municipio" and celda["Estado"] != "Aguascalientes":
                continue
            datos[clave].append([int(celda["Año"]), int(celda["Trimestre"]),
                                 round(float(celda["Indice"]), 2)])

    for clave, puntos in datos.items():
        puntos.sort()
        if not puntos:
            raise SystemExit(f"La serie '{clave}' quedó vacía — ¿cambió el formato del XLSX?")

    ultimo = datos["nacional"][-1]
    corte = f"{ultimo[1]}T{ultimo[0]}"
    salida = {
        "fuente": "Índice SHF de Precios de la Vivienda en México (base 2017=100), "
                  "Sociedad Hipotecaria Federal — datos abiertos (Libre Uso MX)",
        "url": "https://www.gob.mx/shf/documentos/90727",
        "corte": corte,
        "nota": "Serie trimestral. Cada punto es [año, trimestre, índice]. La variación anual "
                "se calcula contra el mismo trimestre del año anterior.",
        "series": [{"clave": clave, "nombre": nombre, "datos": datos[clave]}
                   for _, _, clave, nombre in SERIES],
    }

    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(salida, f, ensure_ascii=False, separators=(",", ":"))
    n = {s["clave"]: len(s["datos"]) for s in salida["series"]}
    print(f"OK → {OUT}  corte={corte}  puntos={n}")


if __name__ == "__main__":
    main()
